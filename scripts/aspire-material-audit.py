#!/usr/bin/env python3
"""Weekly Aspire Material Receipt Audit — Black Hill Landscaping.

Pulls Won Landscape opportunities (DivisionID=17) and flags any where the
estimated material cost has not been entered into actuals yet. Emails Evelin,
Denisse, and Grace (Ops) every Monday 8am Central with a punch list so Grace
can catch up on receipt entry.

A job is flagged when:
  - EstimatedMaterialCost >= $50, AND
  - ActualCostMaterial < 50% of estimate, AND
  - Job is Complete OR PercentComplete > 10%

If no jobs are flagged, no email is sent.

Cron runs at both 13:00 UTC and 14:00 UTC Monday; this script only sends when
local Chicago time hour == 8, giving 8am Central year-round through DST.

Usage:
  python3 aspire-material-audit.py              # Run audit
  python3 aspire-material-audit.py --test       # Test API + email auth
  python3 aspire-material-audit.py --force      # Skip 8am Central guard
  python3 aspire-material-audit.py --dry-run    # Run audit but don't send email
"""

import io, json, os, sys, signal, smtplib, urllib.request, urllib.error, urllib.parse
from datetime import datetime, timezone, timedelta, date
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from email import encoders
from email.utils import formataddr

import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

try:
    from zoneinfo import ZoneInfo
except ImportError:
    from backports.zoneinfo import ZoneInfo

# --- Timeout guard ---
def _timeout_handler(signum, frame):
    log("TIMEOUT: Script exceeded time limit")
    sys.exit(1)

signal.signal(signal.SIGALRM, _timeout_handler)
signal.alarm(120)

# --- Config ---
CONFIG_FILE = os.path.expanduser("~/.config/aspire/config.json")
RECIPIENTS = [
    "evelin@blackhilltx.com",
    # TEST MODE: Denisse + Ops will be added after Evelin reviews the format
    # "denisse@blackhilltx.com",
    # "Ops@blackhilltx.com",
]
ASPIRE_PORTAL = "https://cloud.youraspire.com"
YTD_START = "2026-01-01"   # Report covers Won opps from this date forward
MIN_MATERIAL_EST = 50.00   # Skip material-gap flagging below this
GAP_THRESHOLD = 0.50       # Flag if actual material < 50% of estimate
MIN_PERCENT_COMPLETE = 0.10  # Don't flag in-production jobs <10% done

DRY_RUN = "--dry-run" in sys.argv
FORCE = "--force" in sys.argv
TEST_MODE = "--test" in sys.argv


def log(msg):
    ts = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    print(f"[{ts}] {msg}", flush=True)


# --- Aspire API ---

def load_config():
    """Prefer ASPIRE_REPORTING_* env vars (GH Actions), fall back to config file."""
    client_id = os.environ.get("ASPIRE_REPORTING_CLIENT_ID")
    secret = os.environ.get("ASPIRE_REPORTING_SECRET")
    if client_id and secret:
        return {"api_base_url": "https://cloud-api.youraspire.com",
                "api_client_id": client_id, "api_secret": secret}
    if os.path.exists(CONFIG_FILE):
        with open(CONFIG_FILE) as f:
            cfg = json.load(f)
            if cfg.get("reporting_client_id"):
                return {"api_base_url": cfg.get("api_base_url", "https://cloud-api.youraspire.com"),
                        "api_client_id": cfg["reporting_client_id"],
                        "api_secret": cfg["reporting_secret"]}
            return cfg
    return None


def authenticate(config):
    base = config["api_base_url"]
    data = json.dumps({"ClientId": config["api_client_id"],
                        "Secret": config["api_secret"]}).encode()
    req = urllib.request.Request(f"{base}/Authorization", data=data,
                                 headers={"Content-Type": "application/json"}, method="POST")
    try:
        with urllib.request.urlopen(req) as resp:
            body = json.loads(resp.read().decode())
            return body.get("Token", "")
    except urllib.error.HTTPError as e:
        log(f"Auth failed: {e.code} {e.read().decode() if e.fp else ''}")
        return None


def odata_query(config, token, endpoint, params):
    base = config["api_base_url"]
    qs = urllib.parse.urlencode(params, safe="=&$,'()@ ")
    url = f"{base}/{endpoint}?{qs}"
    req = urllib.request.Request(url, headers={"Authorization": f"Bearer {token}"})
    try:
        with urllib.request.urlopen(req) as resp:
            body = json.loads(resp.read().decode())
            return body.get("value", body) if isinstance(body, dict) else body
    except urllib.error.HTTPError as e:
        log(f"Query failed: {endpoint} - {e.code} {e.read().decode() if e.fp else ''}")
        return []


# --- Vendor hint ---

VENDOR_HINTS = [
    (("sod", "tiftuf", "zoysia", "bermuda"), "Prime Sod / King Ranch"),
    (("dg ", "decomposed granite"), "CFM Stone / SiteOne"),
    (("mulch",), "Organic Recycler"),
    (("edging",), "Ewing (steel edging)"),
    (("drain", "drainage"), "Ewing / SiteOne (drain pipe)"),
    (("lighting", "lights"), "Lighting vendor"),
    (("topdress", "soil",), "Mayer (soil)"),
    (("flower", "annual", "color"), "Annual flower vendor"),
    (("plant", "azalea", "juniper", "tree"), "Plant vendor"),
    (("stone", "boulder", "flagstone"), "CFM / Stone vendor"),
]

def guess_vendor(opp_name):
    name = (opp_name or "").lower()
    hits = []
    for keywords, label in VENDOR_HINTS:
        for kw in keywords:
            if kw in name:
                hits.append(label)
                break
    if hits:
        return " + ".join(dict.fromkeys(hits))  # dedupe while preserving order
    return "Unknown - check job notes"


# --- Audit ---

def fetch_landscape_opps(config, token):
    """Pull all Won Landscape opps for the year with full financial fields."""
    rows = odata_query(config, token, "Opportunities", {
        "$filter": (f"DivisionID eq 17 and OpportunityStatusID eq 15 "
                    f"and WonDate ge {YTD_START}T00:00:00Z"),
        "$select": ("OpportunityID,OpportunityNumber,OpportunityName,PropertyName,"
                    "WonDate,CompleteDate,JobStatusName,PercentComplete,"
                    "EstimatedDollars,EstimatedLaborHours,EstimatedLaborCost,"
                    "EstimatedMaterialCost,ActualLaborHours,ActualCostLabor,"
                    "ActualCostMaterial,ActualCostDollars,ActualEarnedRevenue,"
                    "ActualGrossMarginDollars,ActualGrossMarginPercent,"
                    "SalesRepContactName"),
        "$top": "300",
        "$orderby": "WonDate desc"
    })
    return rows


def shape_row(o):
    """Normalize one Aspire opp into a flat dict with derived fields."""
    status = "Complete" if (o.get("JobStatusName") == "Complete") else "In Production"
    est_lab = o.get("EstimatedLaborCost") or 0
    act_lab = o.get("ActualCostLabor") or 0
    est_mat = o.get("EstimatedMaterialCost") or 0
    act_mat = o.get("ActualCostMaterial") or 0
    est_hrs = o.get("EstimatedLaborHours") or 0
    act_hrs = o.get("ActualLaborHours") or 0
    gm_pct = o.get("ActualGrossMarginPercent")
    profit = o.get("ActualGrossMarginDollars")
    return {
        "opp_id": o.get("OpportunityID"),
        "opp_num": o.get("OpportunityNumber") or o.get("OpportunityID"),
        "name": (o.get("OpportunityName") or "").strip(),
        "property": (o.get("PropertyName") or "").strip(),
        "status": status,
        "won_date": (o.get("WonDate") or "")[:10],
        "complete_date": (o.get("CompleteDate") or "")[:10],
        "pct_complete": o.get("PercentComplete"),
        "est_dollars": o.get("EstimatedDollars") or 0,
        "est_hours": est_hrs,
        "act_hours": act_hrs,
        "hr_variance": (act_hrs - est_hrs) if (status == "Complete") else None,
        "hr_var_pct": ((act_hrs - est_hrs) / est_hrs) if (status == "Complete" and est_hrs) else None,
        "est_labor": est_lab,
        "act_labor": act_lab,
        "labor_var": (act_lab - est_lab) if (status == "Complete") else None,
        "est_material": est_mat,
        "act_material": act_mat,
        "material_var": (act_mat - est_mat) if (status == "Complete" and (act_mat > 0 or est_mat < MIN_MATERIAL_EST)) else None,
        "material_not_logged": (status == "Complete" and act_mat == 0 and est_mat >= MIN_MATERIAL_EST),
        "profit": profit,
        "gm_pct": (gm_pct / 100.0) if gm_pct is not None else None,
        "sales_rep": (o.get("SalesRepContactName") or "").replace(" Montenegro", "").strip(),
        "vendor_hint": guess_vendor(o.get("OpportunityName") or ""),
    }


def root_cause(r):
    if r["status"] != "Complete":
        return ""
    if r["gm_pct"] is None:
        return ""
    if r["gm_pct"] < 0:
        return "LOSS"
    lab_var = r["labor_var"] or 0
    labor_problem = lab_var > 100
    material_problem = (not r["material_not_logged"]) and (r["material_var"] or 0) > 100
    if labor_problem and material_problem:
        return f"BOTH OVER (lab +${lab_var:.0f}, mat +${r['material_var']:.0f})"
    if labor_problem:
        return f"LABOR OVERRUN (+${lab_var:.0f})"
    if material_problem:
        return f"MATERIAL OVERRUN (+${r['material_var']:.0f})"
    if r["material_not_logged"] and r["est_material"] >= 200:
        return f"MATERIAL NOT LOGGED (est ${r['est_material']:.0f})"
    if r["gm_pct"] >= 0.80:
        return "BIG WIN"
    return "ON BUDGET"


def find_gaps(rows):
    """Return rows where material receipts are missing (Complete or >10% in-prod)."""
    flagged = []
    for r in rows:
        est_mat = r["est_material"]
        act_mat = r["act_material"]
        if est_mat < MIN_MATERIAL_EST:
            continue
        if r["status"] != "Complete" and (r["pct_complete"] or 0) < MIN_PERCENT_COMPLETE:
            continue
        if act_mat >= est_mat * GAP_THRESHOLD:
            continue
        flagged.append({**r, "gap": round(est_mat - act_mat, 2)})
    flagged.sort(key=lambda x: (x["status"] != "Complete", -x["gap"]))
    return flagged


# --- Reporting ---

def build_xlsx(all_rows, flagged):
    """Build the full P&L workbook with 5 tabs and return bytes."""
    # Styles
    HEADER = PatternFill(start_color="305496", end_color="305496", fill_type="solid")
    GREEN = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    LIGHT_GREEN = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
    YELLOW = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    RED = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    LIGHT_RED = PatternFill(start_color="F8CBAD", end_color="F8CBAD", fill_type="solid")
    GRAY = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
    TOTAL = PatternFill(start_color="FFD966", end_color="FFD966", fill_type="solid")
    HEADER_FONT = Font(color="FFFFFF", bold=True)
    BOLD = Font(bold=True)
    RED_FONT = Font(color="9C0006", bold=True)
    GREEN_FONT = Font(color="006100", bold=True)
    thin = Side(border_style="thin", color="999999")
    BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)

    wb = openpyxl.Workbook()

    # ===== Sheet 1: Job P&L =====
    ws = wb.active
    ws.title = "Job P&L"
    headers = [
        "Status","Completed Date","Opp #","Sales Rep","Property","Job Name",
        "Est $",
        "Est Hrs","Actual Hrs","Hr Variance","Hr Var %",
        "Est Labor $","Actual Labor $","Labor $ Var",
        "Est Material $","Actual Material $","Material $ Var",
        "Profit/Loss $","GM %","Root Cause"
    ]
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=col, value=h)
        c.fill = HEADER; c.font = HEADER_FONT
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = BORDER
    ws.row_dimensions[1].height = 38
    ws.freeze_panes = "G2"

    # Sort: Complete by complete_date desc, then In Production by won_date desc
    def sort_key(r):
        is_complete = r["status"] == "Complete"
        d = r["complete_date"] if is_complete else r["won_date"]
        return (0 if is_complete else 1, -int(d.replace("-", "") or "0") if d else 0)

    sorted_rows = sorted(all_rows, key=sort_key)

    # Totals tracking
    total_complete = {"est": 0, "earned": 0, "est_hrs": 0, "act_hrs": 0,
                      "est_lab": 0, "act_lab": 0, "est_mat": 0, "act_mat": 0,
                      "profit": 0, "count": 0}
    total_inprog = {"est": 0, "earned": 0, "est_hrs": 0, "act_hrs": 0,
                    "est_lab": 0, "act_lab": 0, "est_mat": 0, "act_mat": 0,
                    "profit": 0, "count": 0}

    for r_idx, r in enumerate(sorted_rows, start=2):
        rc = root_cause(r)
        row_data = [
            r["status"], r["complete_date"] or "", r["opp_num"], r["sales_rep"],
            r["property"], r["name"], round(r["est_dollars"], 2),
            r["est_hours"], r["act_hours"],
            round(r["hr_variance"], 2) if r["hr_variance"] is not None else None,
            r["hr_var_pct"],
            round(r["est_labor"], 2), round(r["act_labor"], 2),
            round(r["labor_var"], 2) if r["labor_var"] is not None else None,
            round(r["est_material"], 2), round(r["act_material"], 2),
            round(r["material_var"], 2) if r["material_var"] is not None else None,
            round(r["profit"], 2) if r["profit"] is not None else None,
            r["gm_pct"], rc,
        ]
        for c_idx, val in enumerate(row_data, start=1):
            cell = ws.cell(row=r_idx, column=c_idx, value=val)
            cell.border = BORDER
            cell.alignment = Alignment(vertical="center",
                                        wrap_text=True if c_idx in (5, 6, 20) else False)

        for col in (7, 12, 13, 14, 15, 16, 17, 18):
            ws.cell(row=r_idx, column=col).number_format = '"$"#,##0.00'
        for col in (8, 9, 10):
            ws.cell(row=r_idx, column=col).number_format = '0.00'
        ws.cell(row=r_idx, column=11).number_format = '0.0%'
        ws.cell(row=r_idx, column=19).number_format = '0.0%'

        # Row fill by GM
        gm = r["gm_pct"]
        if r["status"] != "Complete" or gm is None:
            row_fill = GRAY
        elif gm < 0.50:
            row_fill = RED
        elif gm < 0.70:
            row_fill = YELLOW
        else:
            row_fill = GREEN
        for c_idx in range(1, len(headers) + 1):
            ws.cell(row=r_idx, column=c_idx).fill = row_fill

        # Labor Var (col 14)
        lab_var = r["labor_var"]
        if lab_var is not None:
            cell = ws.cell(row=r_idx, column=14)
            if lab_var > 100:
                cell.fill = RED; cell.font = RED_FONT
            elif lab_var > 25:
                cell.fill = LIGHT_RED
            elif lab_var < -100:
                cell.fill = GREEN; cell.font = GREEN_FONT
            elif lab_var < -25:
                cell.fill = LIGHT_GREEN

        # Material Var (col 17)
        mat_cell = ws.cell(row=r_idx, column=17)
        if r["material_not_logged"]:
            mat_cell.value = "NOT LOGGED"
            mat_cell.fill = GRAY
            mat_cell.font = Font(italic=True, color="595959")
            mat_cell.alignment = Alignment(horizontal="center", vertical="center")
        elif r["material_var"] is not None:
            mv = r["material_var"]
            if mv > 100:
                mat_cell.fill = RED; mat_cell.font = RED_FONT
            elif mv > 25:
                mat_cell.fill = LIGHT_RED
            elif mv < -100:
                mat_cell.fill = GREEN; mat_cell.font = GREEN_FONT
            elif mv < -25:
                mat_cell.fill = LIGHT_GREEN

        # Root Cause (col 20)
        rc_cell = ws.cell(row=r_idx, column=20)
        if rc == "LOSS":
            rc_cell.fill = RED; rc_cell.font = RED_FONT
        elif rc.startswith("LABOR OVERRUN") or rc.startswith("BOTH OVER") or rc.startswith("MATERIAL OVERRUN"):
            rc_cell.fill = LIGHT_RED; rc_cell.font = RED_FONT
        elif rc.startswith("MATERIAL NOT LOGGED"):
            rc_cell.fill = GRAY; rc_cell.font = Font(italic=True, color="595959")
        elif rc == "BIG WIN":
            rc_cell.fill = GREEN; rc_cell.font = GREEN_FONT
        elif rc == "ON BUDGET":
            rc_cell.fill = LIGHT_GREEN

        # GM cell font
        if gm is not None:
            if gm < 0.50:
                ws.cell(row=r_idx, column=19).font = RED_FONT
            elif gm >= 0.70:
                ws.cell(row=r_idx, column=19).font = GREEN_FONT

        # Tally
        bucket = total_complete if r["status"] == "Complete" else total_inprog
        bucket["count"] += 1
        bucket["est"] += r["est_dollars"]
        bucket["est_hrs"] += r["est_hours"]
        bucket["act_hrs"] += r["act_hours"]
        bucket["est_lab"] += r["est_labor"]
        bucket["act_lab"] += r["act_labor"]
        bucket["est_mat"] += r["est_material"]
        bucket["act_mat"] += r["act_material"]
        if r["profit"] is not None:
            bucket["profit"] += r["profit"]

    # Totals rows
    ws.cell(row=ws.max_row + 1, column=1, value="")
    for label, t in (("TOTALS COMPLETED", total_complete), ("TOTALS IN PRODUCTION", total_inprog)):
        rr = ws.max_row + 1
        ws.cell(row=rr, column=1, value=f"{label} ({t['count']} jobs)")
        ws.cell(row=rr, column=7, value=round(t["est"], 2))
        ws.cell(row=rr, column=8, value=round(t["est_hrs"], 2))
        ws.cell(row=rr, column=9, value=round(t["act_hrs"], 2))
        if t["est_hrs"]:
            ws.cell(row=rr, column=10, value=round(t["act_hrs"] - t["est_hrs"], 2))
            ws.cell(row=rr, column=11, value=(t["act_hrs"] - t["est_hrs"]) / t["est_hrs"])
        ws.cell(row=rr, column=12, value=round(t["est_lab"], 2))
        ws.cell(row=rr, column=13, value=round(t["act_lab"], 2))
        ws.cell(row=rr, column=14, value=round(t["act_lab"] - t["est_lab"], 2))
        ws.cell(row=rr, column=15, value=round(t["est_mat"], 2))
        ws.cell(row=rr, column=16, value=round(t["act_mat"], 2))
        ws.cell(row=rr, column=17, value=round(t["act_mat"] - t["est_mat"], 2))
        ws.cell(row=rr, column=18, value=round(t["profit"], 2))
        if t["est"]:
            ws.cell(row=rr, column=19, value=t["profit"] / t["est"])
        for c_idx in range(1, len(headers) + 1):
            ws.cell(row=rr, column=c_idx).fill = TOTAL
            ws.cell(row=rr, column=c_idx).font = BOLD
            ws.cell(row=rr, column=c_idx).border = BORDER
        for col in (7, 12, 13, 14, 15, 16, 17, 18):
            ws.cell(row=rr, column=col).number_format = '"$"#,##0.00'
        for col in (8, 9, 10):
            ws.cell(row=rr, column=col).number_format = '0.00'
        ws.cell(row=rr, column=11).number_format = '0.0%'
        ws.cell(row=rr, column=19).number_format = '0.0%'

    widths = [13, 12, 7, 11, 30, 38, 11, 9, 10, 10, 9, 11, 12, 11, 12, 13, 12, 12, 8, 32]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # ===== Sheet 2: Missing Material Audit =====
    ws2 = wb.create_sheet("Missing Material Audit")
    audit_headers = ["Opp #", "Completed", "Status", "Property", "Job", "Sales Rep",
                     "Est Material $", "Logged $", "Gap $", "Likely Vendor", "Aspire Link"]
    for col, h in enumerate(audit_headers, 1):
        c = ws2.cell(row=1, column=col, value=h)
        c.fill = HEADER; c.font = HEADER_FONT
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = BORDER
    ws2.row_dimensions[1].height = 32
    ws2.freeze_panes = "A2"

    for r_idx, r in enumerate(flagged, start=2):
        opp_url = f"{ASPIRE_PORTAL}/app/opportunities/{r['opp_id']}"
        row_data = [
            r["opp_num"], r["complete_date"] or "", r["status"], r["property"],
            r["name"], r["sales_rep"], r["est_material"], r["act_material"],
            r["gap"], r["vendor_hint"], "Open in Aspire",
        ]
        for c_idx, val in enumerate(row_data, start=1):
            cell = ws2.cell(row=r_idx, column=c_idx, value=val)
            cell.border = BORDER
            cell.alignment = Alignment(vertical="center",
                                        wrap_text=True if c_idx in (4, 5, 10) else False)
        link = ws2.cell(row=r_idx, column=11)
        link.hyperlink = opp_url
        link.font = Font(color="0563C1", underline="single")
        for col in (7, 8, 9):
            ws2.cell(row=r_idx, column=col).number_format = '"$"#,##0.00'
        if r["status"] == "In Production":
            fill = GRAY
        elif r["act_material"] == 0:
            fill = RED
        elif r["gap"] > 100:
            fill = YELLOW
        else:
            fill = GREEN
        for c_idx in range(1, len(audit_headers) + 1):
            ws2.cell(row=r_idx, column=c_idx).fill = fill
        ws2.cell(row=r_idx, column=9).font = BOLD

    total_gap = sum(r["gap"] for r in flagged)
    rr = ws2.max_row + 2
    ws2.cell(row=rr, column=1, value=f"TOTAL ({len(flagged)} jobs)")
    ws2.cell(row=rr, column=7, value=round(sum(r["est_material"] for r in flagged), 2))
    ws2.cell(row=rr, column=8, value=round(sum(r["act_material"] for r in flagged), 2))
    ws2.cell(row=rr, column=9, value=round(total_gap, 2))
    for col in (1, 7, 8, 9):
        ws2.cell(row=rr, column=col).fill = TOTAL
        ws2.cell(row=rr, column=col).font = BOLD
        ws2.cell(row=rr, column=col).border = BORDER
    for col in (7, 8, 9):
        ws2.cell(row=rr, column=col).number_format = '"$"#,##0.00'

    widths2 = [8, 12, 14, 30, 38, 12, 14, 12, 12, 32, 18]
    for i, w in enumerate(widths2, 1):
        ws2.column_dimensions[get_column_letter(i)].width = w

    # ===== Sheet 3: By Vendor (Completed) =====
    vendor_gaps = {}
    for r in flagged:
        if r["status"] != "Complete":
            continue
        v = r["vendor_hint"].split("/")[0].split("+")[0].strip()
        vendor_gaps[v] = vendor_gaps.get(v, 0) + r["gap"]

    ws3 = wb.create_sheet("By Vendor (Completed)")
    ws3["A1"] = "Likely Vendor"
    ws3["B1"] = "Total Gap $ (Completed Only)"
    for col in ("A1", "B1"):
        ws3[col].fill = HEADER; ws3[col].font = HEADER_FONT; ws3[col].border = BORDER
    row_i = 2
    for v, total in sorted(vendor_gaps.items(), key=lambda x: -x[1]):
        ws3.cell(row=row_i, column=1, value=v).border = BORDER
        c = ws3.cell(row=row_i, column=2, value=round(total, 2))
        c.number_format = '"$"#,##0.00'; c.border = BORDER
        row_i += 1
    ws3.column_dimensions["A"].width = 32
    ws3.column_dimensions["B"].width = 30

    # ===== Sheet 4: By Sales Rep =====
    rep_stats = {}
    for r in all_rows:
        if r["status"] != "Complete" or r["gm_pct"] is None:
            continue
        rep = r["sales_rep"] or "Unknown"
        s = rep_stats.setdefault(rep, {"jobs": 0, "est": 0, "earned": 0, "profit": 0,
                                       "under50": 0, "loss_count": 0, "loss_dollars": 0})
        s["jobs"] += 1
        s["est"] += r["est_dollars"]
        s["earned"] += r["est_dollars"]  # earned ≈ est for complete fixed-price
        s["profit"] += r["profit"] or 0
        if r["gm_pct"] < 0.50:
            s["under50"] += 1
        if r["gm_pct"] < 0:
            s["loss_count"] += 1
            s["loss_dollars"] += r["profit"] or 0

    ws4 = wb.create_sheet("By Sales Rep")
    rep_headers = ["Sales Rep", "Jobs Won", "Total Est $", "Total Profit $",
                   "Avg GM %", "# Jobs < 50% GM", "# Losses", "Loss $"]
    for col, h in enumerate(rep_headers, 1):
        c = ws4.cell(row=1, column=col, value=h)
        c.fill = HEADER; c.font = HEADER_FONT
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = BORDER
    ws4.row_dimensions[1].height = 32

    row_i = 2
    for rep, s in sorted(rep_stats.items(), key=lambda x: -x[1]["profit"]):
        avg_gm = s["profit"] / s["earned"] if s["earned"] else 0
        full = (rep, s["jobs"], round(s["est"], 2), round(s["profit"], 2),
                avg_gm, s["under50"], s["loss_count"], round(s["loss_dollars"], 2))
        for c_idx, val in enumerate(full, start=1):
            cell = ws4.cell(row=row_i, column=c_idx, value=val)
            cell.border = BORDER
            cell.alignment = Alignment(vertical="center",
                                        horizontal="center" if c_idx in (2, 6, 7) else "left")
        for col in (3, 4, 8):
            ws4.cell(row=row_i, column=col).number_format = '"$"#,##0.00'
        ws4.cell(row=row_i, column=5).number_format = '0.0%'
        row_i += 1

    widths4 = [22, 11, 16, 16, 11, 16, 11, 14]
    for i, w in enumerate(widths4, 1):
        ws4.column_dimensions[get_column_letter(i)].width = w

    # ===== Sheet 5: Legend =====
    leg = wb.create_sheet("Legend")
    leg["A1"] = "How to Read This Workbook"
    leg["A1"].font = Font(bold=True, size=14)

    leg["A3"] = "Job P&L tab - row colors (overall GM):"; leg["A3"].font = BOLD
    leg["A4"] = "Green: GM >= 70% (healthy profit)"; leg["A4"].fill = GREEN
    leg["A5"] = "Yellow: GM 50-69% (marginal)"; leg["A5"].fill = YELLOW
    leg["A6"] = "Red: GM < 50% (unprofitable or loss)"; leg["A6"].fill = RED
    leg["A7"] = "Gray: In Production (margins not final)"; leg["A7"].fill = GRAY

    leg["A9"] = "Labor $ Var and Material $ Var cells:"; leg["A9"].font = BOLD
    leg["A10"] = "Bright Red: > $100 OVER estimate (problem)"; leg["A10"].fill = RED
    leg["A11"] = "Light Red: $25-$100 over estimate"; leg["A11"].fill = LIGHT_RED
    leg["A12"] = "Light Green: $25-$100 under estimate"; leg["A12"].fill = LIGHT_GREEN
    leg["A13"] = "Bright Green: > $100 UNDER estimate (saved money)"; leg["A13"].fill = GREEN
    leg["A14"] = "'NOT LOGGED' (gray): material was bid but no receipt entered yet"; leg["A14"].fill = GRAY

    leg["A16"] = "Root Cause tags:"; leg["A16"].font = BOLD
    leg["A17"] = "LOSS - job lost money"; leg["A17"].fill = RED
    leg["A18"] = "LABOR OVERRUN +$X - crew burned more hours than bid"; leg["A18"].fill = LIGHT_RED
    leg["A19"] = "MATERIAL OVERRUN +$X - vendor charged more than bid"; leg["A19"].fill = LIGHT_RED
    leg["A20"] = "BOTH OVER - hours AND material both over budget"; leg["A20"].fill = LIGHT_RED
    leg["A21"] = "MATERIAL NOT LOGGED - receipts haven't been entered"; leg["A21"].fill = GRAY
    leg["A22"] = "ON BUDGET - GM 60-80%, no major variance"; leg["A22"].fill = LIGHT_GREEN
    leg["A23"] = "BIG WIN - GM >= 80%"; leg["A23"].fill = GREEN

    leg["A25"] = "Missing Material Audit tab - Grace's action list:"; leg["A25"].font = BOLD
    leg["A26"] = "1. Pull invoices/receipts from each vendor (Prime Sod, King Ranch, Ewing, CFM, Mayer, Organic Recycler, plant vendors)"
    leg["A27"] = "2. Match each receipt to the job by date + property"
    leg["A28"] = f"3. Enter in Aspire: {ASPIRE_PORTAL}/app/purchasing/receipts"
    leg["A29"] = "4. Code each line to the work ticket so the job P&L re-flows correctly"

    leg.column_dimensions["A"].width = 100

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def build_email(flagged):
    today = date.today()
    week_label = today.strftime("%b %d, %Y")
    total_gap = sum(r["gap"] for r in flagged)
    n_complete = sum(1 for r in flagged if r["status"] == "Complete")
    n_inprog = sum(1 for r in flagged if r["status"] != "Complete")

    subject = f"Aspire Material Receipts - {len(flagged)} jobs need entry (${total_gap:,.0f} pending) - {week_label}"

    html = f"""<html><body style="font-family:Arial,sans-serif;color:#222;font-size:14px;">
<h2 style="margin:0 0 8px 0;">Material Receipt Audit - Week of {week_label}</h2>
<p><strong>{len(flagged)}</strong> Landscape jobs are missing material receipts in Aspire.<br>
&nbsp;&nbsp;- {n_complete} completed job(s)<br>
&nbsp;&nbsp;- {n_inprog} in-production job(s) (&gt;10% complete)<br>
<strong>Total material gap: ${total_gap:,.2f}</strong></p>
<p style="background:#FFF3CD;padding:10px;border-left:4px solid #FFC107;">
<strong>Grace</strong>: Punch list is attached as an Excel file. Pull each vendor's receipts/invoices, match them to the job by date + property, then enter via
<a href="{ASPIRE_PORTAL}/app/purchasing/receipts">Purchasing &gt; Purchase Receipts</a> in Aspire.
</p>
<p style="color:#666;font-size:12px;margin-top:20px;">
Audit window: Won Landscape opps since {YTD_START}. Flagged when est material &ge; ${MIN_MATERIAL_EST:.0f} and actual logged &lt; {GAP_THRESHOLD*100:.0f}% of estimate (job Complete or &gt;{MIN_PERCENT_COMPLETE*100:.0f}% in production).
</p>
</body></html>"""

    plain = (
        f"Material Receipt Audit - Week of {week_label}\n"
        f"{len(flagged)} jobs missing material receipts. Total gap: ${total_gap:,.2f}\n\n"
        f"Punch list attached as Excel. Enter receipts at: {ASPIRE_PORTAL}/app/purchasing/receipts"
    )

    return subject, html, plain


def send_email(subject, html_body, plain_body, xlsx_bytes, xlsx_filename):
    sender = os.environ.get("GMAIL_EMAIL")
    password = os.environ.get("GMAIL_APP_PASSWORD")
    if not sender or not password:
        log("No email credentials - printing report to stdout")
        print(f"\nSubject: {subject}\n")
        print(plain_body)
        return False

    msg = MIMEMultipart()
    msg["Subject"] = subject
    msg["From"] = formataddr(("Black Hill Aspire Audit", sender))
    msg["To"] = ", ".join(RECIPIENTS)

    # Body (alternative plain + html)
    body = MIMEMultipart("alternative")
    body.attach(MIMEText(plain_body, "plain"))
    body.attach(MIMEText(html_body, "html"))
    msg.attach(body)

    # XLSX attachment
    part = MIMEBase("application",
                    "vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    part.set_payload(xlsx_bytes)
    encoders.encode_base64(part)
    part.add_header("Content-Disposition", f'attachment; filename="{xlsx_filename}"')
    msg.attach(part)

    try:
        with smtplib.SMTP("smtp.gmail.com", 587) as s:
            s.starttls()
            s.login(sender, password)
            s.sendmail(sender, RECIPIENTS, msg.as_string())
        log(f"Email sent to {len(RECIPIENTS)} recipients with {xlsx_filename} attached")
        return True
    except Exception as e:
        log(f"Email failed: {e}")
        print(f"\nSubject: {subject}\n")
        print(plain_body)
        return False


# --- Main ---

def main():
    log("Aspire Material Audit starting")

    # 8am Central guard (when running on cron with two UTC times)
    if not FORCE and not TEST_MODE:
        central_now = datetime.now(ZoneInfo("America/Chicago"))
        if central_now.weekday() == 0 and central_now.hour != 8:
            log(f"Skipping - current Central hour is {central_now.hour:02d}, only run at 08")
            sys.exit(0)

    config = load_config()
    if not config:
        log("ERROR: No Aspire config found")
        sys.exit(1)

    token = authenticate(config)
    if not token:
        log("ERROR: Aspire auth failed")
        sys.exit(1)
    log("Aspire authenticated")

    if TEST_MODE:
        log("Test mode - API connection OK")
        # Also verify email credentials
        sender = os.environ.get("GMAIL_EMAIL")
        if sender:
            log(f"Email sender configured: {sender}")
        else:
            log("WARNING: GMAIL_EMAIL not set")
        sys.exit(0)

    raw_opps = fetch_landscape_opps(config, token)
    all_rows = [shape_row(o) for o in raw_opps]
    log(f"Fetched {len(all_rows)} Landscape opportunities (YTD)")

    flagged = find_gaps(all_rows)
    log(f"Flagged {len(flagged)} jobs with material gaps")

    if not flagged:
        log("No material gaps - skipping email")
        return

    subject, html, plain = build_email(flagged)
    xlsx_bytes = build_xlsx(all_rows, flagged)
    xlsx_filename = f"BlackHill_Landscape_PnL_{date.today().isoformat()}.xlsx"

    if DRY_RUN:
        log(f"Dry run - would attach {xlsx_filename} ({len(xlsx_bytes):,} bytes)")
        print(f"\nSubject: {subject}\n")
        print(plain)
        with open(f"/tmp/{xlsx_filename}", "wb") as f:
            f.write(xlsx_bytes)
        log(f"Wrote /tmp/{xlsx_filename}")
        return

    send_email(subject, html, plain, xlsx_bytes, xlsx_filename)


if __name__ == "__main__":
    main()
